# -*- coding: utf-8 -*-
"""
POI 최종 엑셀(와이드) -> 대시보드용 집계 JSON 빌드.

원본: 참고자료/■중요■POI_최종(260723).xlsx (시트 POI_탄소발자국)
산출물 (data/):
  - factors.json, meta.json, pois.json, poi_monthly.json

UI/스키마 호환:
  - (kt)관광지_현지인_외지인.xlsx → cont_id별 vL(현지인)·vO(외지인) 집계
  - 소분류 미제공 → scls=중분류
  - cont_id 미제공 → 기존 pois.json 이름·시도·시군구 매칭으로 회수, 없으면 안정 해시 ID
  - 월별 탄소 수식값 미캐시 → 방문자×계수×가중치×EWrt 로 재계산
"""
import openpyxl, json, os, hashlib, re
from datetime import date

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "참고자료", "■중요■POI_최종(260723).xlsx")
REF_DIR = os.path.join(ROOT, "참고자료")
SHEET = "POI_탄소발자국"
OUT = os.path.join(ROOT, "data")
os.makedirs(OUT, exist_ok=True)


def find_kt_nati_src() -> str | None:
    """참고자료/(kt)관광지_현지인_외지인.xlsx 탐색."""
    if not os.path.isdir(REF_DIR):
        return None
    for name in os.listdir(REF_DIR):
        lower = name.lower()
        if (
            name.endswith(".xlsx")
            and not name.startswith("~")
            and "kt" in lower
            and "xlsx" in lower
        ):
            return os.path.join(REF_DIR, name)
    return None


def load_nati_totals(path: str) -> dict[str, dict[str, float]]:
    """cont_id → {vL, vO} (현지인·외지인 누적 방문자)."""
    print("loading nati workbook...", path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = "result" if "result" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    out: dict[str, dict[str, float]] = {}
    rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        cid = str(row[1]).strip() if row[1] is not None else ""
        nati = str(row[3]).strip() if row[3] is not None else ""
        val = num(row[9])
        if not cid:
            continue
        bucket = out.setdefault(cid, {"vL": 0.0, "vO": 0.0})
        if nati == "현지인":
            bucket["vL"] += val
        elif nati == "외지인":
            bucket["vO"] += val
        rows += 1
        if rows % 200000 == 0:
            print("  nati rows:", rows)
    wb.close()
    print("nati rows:", rows, "ids:", len(out))
    return out

with open(os.path.join(ROOT, "scripts", "factors.json"), encoding="utf-8") as f:
    FCONF = json.load(f)
FACTORS = dict(FCONF["factors"])
FDEFAULT = FCONF["default"]
LOW_TH = FCONF["lowCarbonThreshold"]

with open(os.path.join(ROOT, "scripts", "sido_centroids.json"), encoding="utf-8") as f:
    CENTROIDS = json.load(f)

COORDS_PATH = os.path.join(OUT, "poi_coords.json")
OLD_POIS_PATH = os.path.join(OUT, "pois.json")

# 와이드 컬럼 인덱스
COL_SERIAL, COL_NM, COL_SIDO, COL_SGG, COL_LCLS, COL_MCLS = 0, 1, 2, 3, 4, 5
COL_VIS_START, COL_VIS_END = 6, 46          # 2023.01 ~ 2026.04
COL_COEF, COL_WEIGHT, COL_EWRT23, COL_EWRT24 = 46, 47, 48, 49


def load_poi_coords_cache():
    if not os.path.isfile(COORDS_PATH):
        return {}
    with open(COORDS_PATH, encoding="utf-8") as f:
        return json.load(f).get("coords", {})


def load_legacy_id_map():
    """기존 pois.json: (sido, sgg, nm) / (sido, nm) → cont_id."""
    by_full = {}
    by_sido_nm = {}
    if not os.path.isfile(OLD_POIS_PATH):
        return by_full, by_sido_nm
    with open(OLD_POIS_PATH, encoding="utf-8") as f:
        for p in json.load(f):
            key = (p.get("sido") or "", p.get("sgg") or "", p.get("nm") or "")
            by_full[key] = str(p["id"])
            by_sido_nm.setdefault((key[0], key[2]), str(p["id"]))
    return by_full, by_sido_nm


POI_COORDS = load_poi_coords_cache()
LEGACY_FULL, LEGACY_SIDO_NM = load_legacy_id_map()


def ym_from_label(label) -> str | None:
    """'2023.01' / datetime → '202301'."""
    if label is None:
        return None
    if hasattr(label, "strftime"):
        return label.strftime("%Y%m")
    s = str(label).strip()
    m = re.match(r"^(\d{4})[.\-/]?(\d{1,2})$", s)
    if not m:
        return None
    return f"{m.group(1)}{int(m.group(2)):02d}"


def stable_id(sido, sgg, nm, serial) -> str:
    raw = f"{sido}|{sgg}|{nm}|{serial}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()[:12]


def resolve_id(sido, sgg, nm, serial) -> str:
    cid = LEGACY_FULL.get((sido or "", sgg or "", nm or ""))
    if cid:
        return cid
    cid = LEGACY_SIDO_NM.get((sido or "", nm or ""))
    if cid:
        return cid
    return stable_id(sido, sgg, nm, serial)


def jitter_coord(sido, cont_id):
    base = CENTROIDS.get(sido, [127.7, 36.5])
    h = hashlib.md5((sido + str(cont_id)).encode("utf-8")).hexdigest()
    a = int(h[0:8], 16) / 0xFFFFFFFF
    b = int(h[8:16], 16) / 0xFFFFFFFF
    spread = 0.55
    return [round(base[0] + (a - 0.5) * spread, 5),
            round(base[1] + (b - 0.5) * spread, 5)]


def resolve_coord(sido, cont_id):
    entry = POI_COORDS.get(str(cont_id))
    if entry and entry.get("source") == "kto":
        lon, lat = entry.get("lon"), entry.get("lat")
        if lon is not None and lat is not None:
            return round(float(lon), 6), round(float(lat), 6)
    lon, lat = jitter_coord(sido, cont_id)
    return lon, lat


def num(v, default=0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def ewrt_for_ym(ym: str, ewrt23: float, ewrt24: float) -> float:
    year = int(ym[:4])
    return ewrt23 if year <= 2023 else ewrt24


print("loading workbook...", SRC)
KT_SRC = find_kt_nati_src()
NATI_TOTALS = load_nati_totals(KT_SRC) if KT_SRC else {}
if not NATI_TOTALS:
    print("WARN: KT 현지인·외지인 파일 없음 → vO=0 fallback")

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
if SHEET not in wb.sheetnames:
    raise SystemExit(f"시트 '{SHEET}' 없음. sheets={wb.sheetnames}")
ws = wb[SHEET]

it = ws.iter_rows(values_only=True)
header0 = next(it)
header1 = next(it)

ym_list = []
for i in range(COL_VIS_START, COL_VIS_END):
    ym = ym_from_label(header1[i] if header1 else None)
    if not ym:
        raise SystemExit(f"월 라벨 파싱 실패 col={i} value={header1[i]!r}")
    ym_list.append(ym)

print("months:", ym_list[0], "~", ym_list[-1], f"({len(ym_list)})")

sido_set = set()
sgg_by_sido = {}
lcls_set = set()
mcls_by_lcls = {}
mcls_coef = {}  # 중분류 → 엑셀 계수(B안) 수집

pois = {}
poi_monthly = {}
legacy_hit = 0
legacy_miss = 0
cnt = 0

for row in it:
    nm = row[COL_NM]
    if nm is None or str(nm).strip() == "":
        continue
    serial = row[COL_SERIAL]
    sido = row[COL_SIDO] or ""
    sgg = row[COL_SGG] or ""
    lcls = row[COL_LCLS] or ""
    mcls = row[COL_MCLS] or ""
    coef = num(row[COL_COEF], FDEFAULT)
    weight = num(row[COL_WEIGHT], 1.0)
    ewrt23 = num(row[COL_EWRT23], 1.0)
    ewrt24 = num(row[COL_EWRT24], 1.0)

    cont_id = resolve_id(sido, sgg, nm, serial)
    matched_legacy = (sido, sgg, nm) in LEGACY_FULL or (sido, nm) in LEGACY_SIDO_NM
    if matched_legacy:
        legacy_hit += 1
    else:
        legacy_miss += 1
    # 동일 이름·지역 중복 행 → 고유 ID 보장
    if cont_id in pois:
        cont_id = f"{cont_id}_{serial}"
        legacy_miss += 1
        legacy_hit = max(0, legacy_hit - 1)

    month_v = {}
    total_v = 0.0
    total_e_kg = 0.0
    for i, ym in enumerate(ym_list):
        v = num(row[COL_VIS_START + i])
        month_v[ym] = v
        total_v += v
        total_e_kg += v * coef * weight * ewrt_for_ym(ym, ewrt23, ewrt24)

    # UI pc: 실효 1인당 kg (총배출kg / 총방문자)
    pc = (total_e_kg / total_v) if total_v > 0 else round(coef * weight, 3)

    nati = NATI_TOTALS.get(cont_id)
    if nati:
        kt_total = nati["vL"] + nati["vO"]
        if kt_total > 0 and total_v > 0:
            scale = total_v / kt_total
            vL = round(nati["vL"] * scale, 1)
            vO = round(nati["vO"] * scale, 1)
        elif kt_total > 0:
            vL = round(nati["vL"], 1)
            vO = round(nati["vO"], 1)
        else:
            vL = round(total_v, 1)
            vO = 0.0
    else:
        vL = round(total_v, 1)
        vO = 0.0

    p = {
        "id": cont_id,
        "nm": str(nm).strip(),
        "sido": sido,
        "sgg": sgg,
        "lcls": lcls,
        "mcls": mcls,
        "scls": mcls,  # 소분류 없음 → 중분류로 채움 (UI 유지)
        "v": round(total_v, 1),
        "vL": vL,
        "vO": vO,
        "e": round(total_e_kg / 1000.0, 2),
        "pc": round(pc, 3),
    }
    pois[cont_id] = p
    poi_monthly[cont_id] = month_v

    sido_set.add(sido)
    sgg_by_sido.setdefault(sido, set()).add(sgg)
    lcls_set.add(lcls)
    mcls_by_lcls.setdefault(lcls, set()).add(mcls)
    if mcls:
        mcls_coef[mcls] = coef

    cnt += 1
    if cnt % 2000 == 0:
        print("  rows:", cnt)

wb.close()
print("rows total:", cnt, "pois:", len(pois))
print(f"legacy cont_id match: hit={legacy_hit} miss={legacy_miss}")
kto_coords = sum(1 for c in POI_COORDS.values() if c.get("source") == "kto")
print(f"coord cache: kto={kto_coords} (poi_coords.json)")

# factors.json: 엑셀 B안 계수로 갱신 (method 페이지용)
for mcls, coef in mcls_coef.items():
    FACTORS[mcls] = round(coef, 3)

# ---- 좌표 ----
poi_list = []
poi_monthly_out = {}
for cid, p in pois.items():
    p["lon"], p["lat"] = resolve_coord(p["sido"], cid)
    poi_list.append(p)
    poi_monthly_out[cid] = [round(poi_monthly[cid].get(y, 0.0), 1) for y in ym_list]

# ---- 전국 월별 추이 ----
nat_month_v = {y: 0.0 for y in ym_list}
nat_month_e = {y: 0.0 for y in ym_list}
for cid, p in pois.items():
    for y in ym_list:
        v = poi_monthly[cid].get(y, 0.0)
        nat_month_v[y] += v
        # 월별 배출: pc 평균 사용 (POI별 실효계수)
        nat_month_e[y] += v * p["pc"] / 1000.0
national_monthly = [
    {"ym": y, "visitors": round(nat_month_v[y], 0), "emission": round(nat_month_e[y], 1)}
    for y in ym_list
]

# ---- 카테고리 롤업 ----
lcls_roll = {}
mcls_roll = {}
for p in poi_list:
    l = lcls_roll.setdefault(p["lcls"], {"lcls": p["lcls"], "visitors": 0.0, "emission": 0.0, "nPoi": 0})
    l["visitors"] += p["v"]; l["emission"] += p["e"]; l["nPoi"] += 1
    m = mcls_roll.setdefault(p["mcls"], {"mcls": p["mcls"], "lcls": p["lcls"], "visitors": 0.0, "emission": 0.0, "nPoi": 0})
    m["visitors"] += p["v"]; m["emission"] += p["e"]; m["nPoi"] += 1

total_e = sum(p["e"] for p in poi_list)
total_v = sum(p["v"] for p in poi_list)

lcls_rollup = sorted(lcls_roll.values(), key=lambda x: -x["emission"])
for x in lcls_rollup:
    x["emission"] = round(x["emission"], 1)
    x["visitors"] = round(x["visitors"], 0)
    x["share"] = round(x["emission"] / total_e * 100, 1) if total_e else 0
mcls_rollup = sorted(mcls_roll.values(), key=lambda x: -x["emission"])
for x in mcls_rollup:
    x["emission"] = round(x["emission"], 1)
    x["visitors"] = round(x["visitors"], 0)
    x["share"] = round(x["emission"] / total_e * 100, 1) if total_e else 0

# ---- 지역 롤업 ----
sido_roll = {}
sgg_roll = {}
for p in poi_list:
    s = sido_roll.setdefault(p["sido"], {"sido": p["sido"], "visitors": 0.0, "emission": 0.0, "nPoi": 0})
    s["visitors"] += p["v"]; s["emission"] += p["e"]; s["nPoi"] += 1
    key = p["sido"] + "|" + p["sgg"]
    g = sgg_roll.setdefault(key, {"sido": p["sido"], "sgg": p["sgg"], "visitors": 0.0, "emission": 0.0, "nPoi": 0})
    g["visitors"] += p["v"]; g["emission"] += p["e"]; g["nPoi"] += 1
for x in sido_roll.values():
    x["emission"] = round(x["emission"], 1); x["visitors"] = round(x["visitors"], 0)
    x["lon"], x["lat"] = CENTROIDS.get(x["sido"], [127.7, 36.5])
for x in sgg_roll.values():
    x["emission"] = round(x["emission"], 1); x["visitors"] = round(x["visitors"], 0)
sido_rollup = sorted(sido_roll.values(), key=lambda x: -x["emission"])
sgg_rollup = sorted(sgg_roll.values(), key=lambda x: -x["emission"])

# ---- KPI ----
n_months = len(ym_list)
total_e_kg = total_e * 1000.0
low_carbon = [p for p in poi_list if p["pc"] <= LOW_TH]
top10 = sorted(poi_list, key=lambda x: -x["e"])[:10]
top10_e = sum(p["e"] for p in top10)

kpis = {
    "nPoi": len(poi_list),
    "totalVisitors": round(total_v, 0),
    "totalEmission": round(total_e, 0),
    "perPoiAvgKg": round(total_e_kg / len(poi_list) / n_months, 1) if poi_list and n_months else 0,
    "perCapitaKg": round(total_e_kg / total_v, 2) if total_v else 0,
    "top10Share": round(top10_e / total_e * 100, 1) if total_e else 0,
    "lowCarbonCount": len(low_carbon),
    "nSido": len(sido_set),
    "nSgg": len(sgg_roll),
}

meta = {
    "ymList": ym_list,
    "ymMin": ym_list[0], "ymMax": ym_list[-1], "nMonths": n_months,
    "updatedAt": date.today().isoformat(),
    "filters": {
        "sido": sorted(sido_set),
        "sggBySido": {k: sorted(v) for k, v in sgg_by_sido.items()},
        "lcls": sorted(lcls_set),
        "mclsByLcls": {k: sorted(v) for k, v in mcls_by_lcls.items()},
        "nati": ["전체", "현지인", "외지인"],
    },
    "kpis": kpis,
    "national_monthly": national_monthly,
    "lclsRollup": lcls_rollup,
    "mclsRollup": mcls_rollup,
    "sidoRollup": sido_rollup,
    "sggRollup": sgg_rollup,
    "top10": [{"id": p["id"], "nm": p["nm"], "sido": p["sido"], "sgg": p["sgg"],
               "visitors": p["v"], "emission": p["e"]} for p in top10],
}

pois_out = [{
    "id": p["id"], "nm": p["nm"], "sido": p["sido"], "sgg": p["sgg"],
    "lcls": p["lcls"], "mcls": p["mcls"], "scls": p["scls"],
    "v": p["v"], "vL": p["vL"], "vO": p["vO"], "e": p["e"], "pc": p["pc"],
    "lon": p["lon"], "lat": p["lat"],
} for p in sorted(poi_list, key=lambda x: -x["e"])]


def dump(name, obj):
    path = os.path.join(OUT, name)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, separators=(",", ":"))
    print("wrote", name, round(os.path.getsize(path) / 1024 / 1024, 2), "MB")


dump("factors.json", {
    "factors": FACTORS,
    "default": FDEFAULT,
    "lowCarbonThreshold": LOW_TH,
    "_comment": "POI_최종(260723) B안 계수. 실제 배출=방문자×계수×업종가중치×EWrt",
})
dump("meta.json", meta)
dump("pois.json", pois_out)
dump("poi_monthly.json", poi_monthly_out)
print("KPIs:", json.dumps(kpis, ensure_ascii=False))
print("DONE")
